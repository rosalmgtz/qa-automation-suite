import pytest
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import re
import os

# --- CONFIGURACIÓN INICIAL Y RUTAS ---
# 🕒 Fecha actual para nombres de archivos
fecha_actual = datetime.now().strftime("%Y-%m-%d")

# Define la ruta de la carpeta reports/ en relación con este script
# os.path.abspath(__file__) es la ruta completa de este archivo.
# os.path.dirname(...) obtiene el directorio donde está (ej. /project/tests).
# '..' sube un nivel (ej. a /project/).
# 'reports' baja a la carpeta reports/ (ej. /project/reports/).
reports_dir = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), '..', 'reports')

# Asegura que la carpeta 'reports' exista. Si no, la crea.
# Esto es crucial para que el Excel se guarde correctamente.
os.makedirs(reports_dir, exist_ok=True)

# Nombre del archivo Excel, guardado directamente en la carpeta reports/
nombre_excel_final = os.path.join(
    reports_dir, f"busquedas_google_SABRINA_{fecha_actual}.xlsx")

# Inicializa el libro de Excel (Workbook) y la hoja de resumen
wb = Workbook()
ws_resumen = wb.active
ws_resumen.title = "Resumen"
ws_resumen.append(["Término", "Título", "Enlace"])

# 🔍 Términos de búsqueda
busquedas = [
    "automatización de pruebas con IA",
    "herramientas de testing con Selenium",
    "evitar CAPTCHA en Selenium"  # Puedes añadir o quitar términos
]
# --- FIN CONFIGURACIÓN INICIAL Y RUTAS ---


# ✨ Limpieza de texto
def limpiar(texto):
    texto = re.sub(r'[^\x00-\x7F]+', '', texto)  # Remueve caracteres no-ASCII
    return texto.strip()


# 🎨 Aplicar estilos a hoja Excel
def aplicar_estilos(ws):
    for cell in ws[1]:  # Aplica estilos a la primera fila (encabezados)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD",
                                end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Aplica colores alternos a las filas
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7",
                           fill_type="solid") if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill

    for col in ws.columns:  # Ajusta el ancho de las columnas automáticamente
        max_length = 0
        column = []
        for cell in col:
            column.append(str(cell.value))
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Un poco de padding
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width


# 🧪 Configurar navegador (Fixture de Pytest)
@pytest.fixture(scope="module")
def driver():
    # Obtiene el valor de la variable de entorno BROWSER_HEADLESS.
    # Por defecto, si no está definida, asume 'true' (modo sin cabeza).
    # Esto permite controlar el modo headless desde el workflow de GitHub Actions.
    headless_mode = os.getenv("BROWSER_HEADLESS", "true").lower() == "true"
    print(f"🟢 Iniciando navegador stealth (headless={headless_mode})...")

    # Configuración de opciones de Chrome para mayor robustez en CI
    options = uc.ChromeOptions()
    # Necesario para entornos Linux como GitHub Actions
    options.add_argument('--no-sandbox')
    # Previene problemas de memoria en CI
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')  # Recomendado para headless
    # Asegura un tamaño de ventana consistente
    options.add_argument('--window-size=1920,1080')
    # Asegura que la ventana se maximice al inicio
    options.add_argument('--start-maximized')
    options.add_argument('--disable-extensions')  # Deshabilita extensiones
    # Deshabilita barras de información
    options.add_argument('--disable-infobars')
    # Indica que el navegador está siendo automatizado
    options.add_argument('--enable-automation')
    # Puede ayudar con TimeoutExceptions
    options.add_argument('--disable-browser-side-navigation')
    # Deshabilita la comprobación del navegador predeterminado
    options.add_argument('--no-default-browser-check')
    # Deshabilita la ejecución por primera vez
    options.add_argument('--no-first-run')
    # Deshabilita las aplicaciones predeterminadas
    options.add_argument('--disable-default-apps')
    # Deshabilita el bloqueo de pop-ups
    options.add_argument('--disable-popup-blocking')

    d = uc.Chrome(headless=headless_mode, use_subprocess=True,
                  options=options)  # Asegura que 'options' se pasa aquí
    # d.maximize_window() # Ya se maneja con --start-maximized en las opciones
    yield d  # Cede el control a los tests

    # Después de que los tests terminen, cierra el navegador
    try:
        d.quit()
    except Exception as e:
        print(f"⚠️ Error al cerrar el navegador: {e}")
    print("🛑 Navegador cerrado.")


# 📘 Personalizar metadata del reporte HTML de Pytest
def pytest_configure(config):
    # Asegúrate que sea el nombre deseado
    config._metadata["Autor"] = "SABRINA"
    config._metadata["Proyecto"] = "Scraping con Selenium"
    config._metadata["Fecha"] = fecha_actual
    config._metadata["Ubicación"] = "Buenavista, México"


# 🧪 Test por cada término de búsqueda (Parametrizado con Pytest)
@pytest.mark.parametrize("query", busquedas)
def test_busqueda_google(driver, query):
    print(f"\n🔎 Buscando: {query}")
    driver.get("https://www.google.com")

    # Aumentar tiempo de espera para el campo de búsqueda
    WebDriverWait(driver, 20).until(  # Aumentado a 20 segundos
        EC.presence_of_element_located((By.NAME, "q"))
    )

    # Consentimiento de cookies / CAPTCHA bypass básico
    try:
        # Intenta aceptar cookies buscando botones comunes
        aceptar_button = WebDriverWait(driver, 15).until(  # Aumentado a 15 segundos
            EC.element_to_be_clickable(
                (By.XPATH, "//button[contains(., 'Aceptar')] | //button[contains(., 'I agree')] | //button[contains(., 'Accept all')]"))
        )
        aceptar_button.click()
        print("🍪 Botón de aceptar cookies clickeado.")
        time.sleep(3)  # Un poco más de tiempo después de clickear
    except:
        # Si no encuentra el botón o falla, intentar cerrar modales con JS
        try:
            driver.execute_script(
                "document.querySelectorAll('iframe, dialog, [aria-modal=\"true\"], [role=\"dialog\"]').forEach(e => e.remove());"
            )
            print("👁️ Modales/iframes removidos con JavaScript.")
            time.sleep(2)
        except:
            print("🚫 No se encontró botón de aceptar cookies ni modales para remover.")
            pass  # Continúa si no hay cookies o modal

    caja = driver.find_element(By.NAME, "q")
    caja.send_keys(query)
    caja.send_keys(Keys.RETURN)

    # Aumentar tiempo de espera para los resultados de búsqueda
    # A veces el ID "search" no es el más robusto, podríamos usar un selector CSS más general
    # o esperar a que un elemento de resultado específico aparezca.
    try:
        WebDriverWait(driver, 20).until(  # Aumentado a 20 segundos
            EC.presence_of_element_located((By.ID, "search"))
        )
        print("DEBUG: Elemento 'search' encontrado.")  # DEBUG LINE
    except Exception as e:
        # DEBUG LINE
        print(
            f"DEBUG: Falló la espera por ID 'search': {e}. Capturando HTML para depuración.")
        # Capturar el HTML de la página para depuración
        with open(os.path.join(reports_dir, f"debug_html_search_fail_{query[:15]}_{fecha_actual}.html"), "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        # Intentar esperar por un selector alternativo si el ID "search" falla
        try:
            WebDriverWait(driver, 10).until(
                # 'rso' es el contenedor principal de resultados
                EC.presence_of_element_located((By.CSS_SELECTOR, "div#rso"))
            )
            print("DEBUG: Elemento 'rso' (alternativo) encontrado.")  # DEBUG LINE
        except Exception as e_alt:
            # DEBUG LINE
            print(f"DEBUG: Falló la espera por 'rso' también: {e_alt}.")
            # Si ambos fallan, re-raise el error para que la prueba falle.
            # Re-lanza el error para que la prueba falle si no encuentra nada.
            raise e_alt

    time.sleep(5)  # Un tiempo extra para que la página cargue completamente

    bloques = driver.find_elements(
        By.CSS_SELECTOR, "div#search .tF2Cxc, div.g")

    # DEBUG LINE
    print(
        f"DEBUG: Número de bloques de resultados encontrados: {len(bloques)}")
    if len(bloques) == 0:
        print(
            f"⚠️ No se encontraron bloques de resultados estándar para '{query}'. Reintentando con selectores alternativos.")
        # Podrías añadir aquí otra capa de lógica o reintentos si es necesario

    assert len(
        bloques) > 0, f"No se encontraron resultados visibles para '{query}' después de la búsqueda."

    ws = wb.create_sheet(title=query[:31])
    ws.append(["Título", "Enlace"])
    count = 0

    for bloque in bloques:
        try:
            titulo_elem = bloque.find_element(By.TAG_NAME, "h3")
            enlace_elem = bloque.find_element(By.TAG_NAME, "a")

            titulo = limpiar(titulo_elem.text)
            enlace = enlace_elem.get_attribute("href").strip()

            if len(titulo) > 5 and enlace and enlace.startswith("http"):
                ws.append([titulo, enlace])
                ws_resumen.append([query, titulo, enlace])
                count += 1
            if count >= 5:
                break
        except Exception as e:
            # DEBUG LINE
            print(f"DEBUG: Error al procesar bloque de resultado: {e}")
            continue

    aplicar_estilos(ws)
    print(f"📋 {count} resultados guardados en hoja: '{query}'")


# 🧾 Guardar Excel al final de todas las pruebas (Hook de Pytest)
def pytest_sessionfinish(session, exitstatus):
    print("DEBUG: pytest_sessionfinish iniciado.")  # DEBUG LINE
    # Remueve la hoja "Sheet" por defecto si está vacía
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 0:
        del wb["Sheet"]

    aplicar_estilos(ws_resumen)

    try:
        wb.save(nombre_excel_final)
        print(f"\n✅ Archivo Excel generado: {nombre_excel_final}")
    except Exception as e:
        # DEBUG LINE
        print(
            f"❌ ERROR: Falló al guardar el archivo Excel '{nombre_excel_final}': {e}")
