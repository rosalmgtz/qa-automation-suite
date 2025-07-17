from openpyxl import Workbook
import re
import time
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc
import subprocess
import sys
import importlib.util

# 🔍 Asegurar openpyxl


def asegurar_openpyxl():
    if importlib.util.find_spec("openpyxl") is None:
        print("📦 Instalando 'openpyxl'...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl"])
    else:
        print("✅ 'openpyxl' ya está disponible.")


asegurar_openpyxl()


class GoogleScraperMulti:
    def __init__(self, queries, max_resultados=5):
        self.queries = queries
        self.max_resultados = max_resultados
        self.driver = None
        fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.output_file = f"busquedas_google_{fecha_actual}.xlsx"
        self.wb = Workbook()

    def iniciar_navegador(self):
        print("🟢 Iniciando navegador stealth...")
        self.driver = uc.Chrome(headless=False, use_subprocess=True)
        self.driver.maximize_window()

    def realizar_busqueda(self, query):
        print(f"\n🔎 Buscando: {query}")
        self.driver.get("https://www.google.com")
        time.sleep(2)
        caja_busqueda = self.driver.find_element(By.NAME, "q")
        caja_busqueda.send_keys(query)
        time.sleep(1)
        caja_busqueda.send_keys(Keys.RETURN)
        time.sleep(3)

    def limpiar_texto(self, texto):
        texto = re.sub(r'[^\x00-\x7F]+', '', texto)
        return texto.strip()

    def extraer_resultados(self):
        bloques = self.driver.find_elements(
            By.CSS_SELECTOR, "div#search .tF2Cxc")
        resultados = []
        count = 0

        for bloque in bloques:
            try:
                titulo = bloque.find_element(By.CSS_SELECTOR, "h3").text
                enlace = bloque.find_element(
                    By.CSS_SELECTOR, "a").get_attribute("href").strip()
                titulo = self.limpiar_texto(titulo)

                if len(titulo) > 5 and titulo.isprintable() and enlace:
                    resultados.append([titulo, enlace])
                    count += 1

                if count >= self.max_resultados:
                    break
            except:
                continue
        print(f"📋 {len(resultados)} resultados extraídos.")
        return resultados

    def guardar_hoja(self, nombre_hoja, datos):
        ws = self.wb.create_sheet(title=nombre_hoja[:31])
        ws.append(["Título", "Enlace"])
        for fila in datos:
            ws.append(fila)

    def cerrar_navegador(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception as e:
                print(f"⚠️ Error al cerrar el navegador: {e}")
            self.driver = None  # Evita que el destructor intente cerrarlo otra vez
            print("🛑 Navegador cerrado.")

    def ejecutar(self):
        try:
            self.iniciar_navegador()
            for query in self.queries:
                self.realizar_busqueda(query)
                resultados = self.extraer_resultados()
                self.guardar_hoja(nombre_hoja=query, datos=resultados)
            if "Sheet" in self.wb.sheetnames and self.wb["Sheet"].max_row == 0:
                del self.wb["Sheet"]
            self.wb.save(self.output_file)
            print(f"\n✅ Archivo Excel generado: {self.output_file}")
        except Exception as e:
            print(f"❌ Error: {e}")
        finally:
            self.cerrar_navegador()


# 🔁 Lista de búsquedas
busquedas = [
    "automatización de pruebas con IA",
    "herramientas de testing con Selenium",
    "aplicaciones móviles con React Native",
    "mapas en apps Android",
    "evitar CAPTCHA en Selenium"
]

# 🚀 Ejecución
if __name__ == "__main__":
    scraper = GoogleScraperMulti(queries=busquedas)
    scraper.ejecutar()
