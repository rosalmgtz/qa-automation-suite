import pytest
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import re

# 🕒 Fecha actual
fecha_actual = datetime.now().strftime("%Y-%m-%d")
nombre_excel = f"busquedas_google_SABRINA_{fecha_actual}.xlsx"
wb = Workbook()

# 🔍 Términos de búsqueda
busquedas = [
    "automatización de pruebas con IA",
    "herramientas de testing con Selenium",
    "aplicaciones móviles con React Native",
    "mapas en apps Android",
    "evitar CAPTCHA en Selenium"
]

# ✨ Limpieza de texto


def limpiar(texto):
    texto = re.sub(r'[^\x00-\x7F]+', '', texto)
    return texto.strip()

# 🎨 Aplicar estilos a hoja Excel


def aplicar_estilos(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD",
                                end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7",
                           fill_type="solid") if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 5

# 🧪 Configurar navegador


@pytest.fixture(scope="module")
def driver():
    print("🟢 Iniciando navegador stealth...")
    d = uc.Chrome(headless=True, use_subprocess=True)
    d.maximize_window()
    yield d
    try:
        d.quit()
    except:
        pass
    print("🛑 Navegador cerrado.")

# 📘 Personalizar metadata del reporte


def pytest_configure(config):
    config._metadata["Autor"] = "SABRINA"
    config._metadata["Proyecto"] = "Scraping con Selenium"
    config._metadata["Fecha"] = fecha_actual
    config._metadata["Ubicación"] = "Buenavista, México"

# 🧪 Test por cada término de búsqueda


@pytest.mark.parametrize("query", busquedas)
def test_busqueda_google(driver, query):
    print(f"\n🔎 Buscando: {query}")
    driver.get("https://www.google.com")
    time.sleep(2)

    try:
        aceptar = driver.find_element(
            By.XPATH, "//button[contains(., 'Aceptar')]")
        aceptar.click()
        time.sleep(2)
    except:
        pass

    caja = driver.find_element(By.NAME, "q")
    caja.send_keys(query)
    caja.send_keys(Keys.RETURN)
    time.sleep(3)

    bloques = driver.find_elements(By.CSS_SELECTOR, "div#search .tF2Cxc")
    if len(bloques) == 0:
        bloques = driver.find_elements(By.CSS_SELECTOR, "div.g")
    assert len(bloques) > 0, f"No se encontraron resultados para '{query}'"

    ws = wb.create_sheet(title=query[:31])
    ws.append(["Título", "Enlace"])
    count = 0

    for bloque in bloques:
        try:
            titulo = limpiar(bloque.find_element(By.TAG_NAME, "h3").text)
            enlace = bloque.find_element(
                By.TAG_NAME, "a").get_attribute("href").strip()
            if len(titulo) > 5 and enlace:
                ws.append([titulo, enlace])
                count += 1
            if count >= 5:
                break
        except:
            continue

    aplicar_estilos(ws)
    print(f"📋 {count} resultados guardados en hoja: {query}")

# 🧾 Guardar Excel al final


def pytest_sessionfinish(session, exitstatus):
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 0:
        del wb["Sheet"]
    wb.save(nombre_excel)
    print(f"\n✅ Archivo Excel generado: {nombre_excel}")
